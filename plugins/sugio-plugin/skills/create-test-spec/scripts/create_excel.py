#!/usr/bin/env python3
"""テスト仕様書のJSONデータからフォーマット済みExcelファイルを生成する。

Usage:
    python create_excel.py <input.json> <output.xlsx>

入力JSONフォーマット:
    {
        "title": "変更概要タイトル",
        "branch": "feature/xxx",
        "sections": [
            {
                "name": "セクション名　テスト種別",
                "cases": [
                    {
                        "no": 1,
                        "category": "分類タグ",
                        "test_item": "テスト項目",
                        "procedure": "操作手順",
                        "expected": "期待結果",
                        "notes": "備考"
                    }
                ]
            }
        ]
    }
"""

import json
import subprocess
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[INFO] openpyxl をインストールしています...", file=sys.stderr)
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 定数
# ---------------------------------------------------------------------------
COLUMNS = ["No", "分類", "テスト項目", "操作手順", "期待結果", "確認者１", "確認者２", "備考"]
COL_WIDTHS = [6, 18, 38, 62, 52, 10, 10, 30]
FONT_NAME = "Yu Gothic"
FONT_SIZE = 10
# 1行あたりの行高（pt）。フォントサイズ10に対して余裕を持たせた値。
LINE_HEIGHT = 15.0
# Excelが受け付ける行高の上限（pt）。
MAX_ROW_HEIGHT = 409.0

# ---------------------------------------------------------------------------
# スタイル定義
# ---------------------------------------------------------------------------
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

STYLES = {
    "title": {
        "font": Font(name=FONT_NAME, bold=True, size=14),
    },
    "branch": {
        "font": Font(name=FONT_NAME, size=11, color="555555"),
    },
    "header": {
        "font": Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11),
        "fill": PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    },
    "section": {
        "font": Font(name=FONT_NAME, bold=True, size=11, color="1F4E79"),
        "fill": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    },
    "data": {
        "font": Font(name=FONT_NAME, size=FONT_SIZE),
        "alignment": Alignment(vertical="top", wrap_text=True),
    },
    "data_even": {
        "fill": PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid"),
    },
    "bug_fix": {
        "font": Font(name=FONT_NAME, size=10, color="CC0000"),
    },
    "new_feature": {
        "font": Font(name=FONT_NAME, size=10, color="007A33"),
    },
}


# ---------------------------------------------------------------------------
# メイン処理
# ---------------------------------------------------------------------------
def count_wrapped_lines(text: str, col_width: int) -> int:
    """折り返しを考慮した表示行数を計算する。

    Excelのカラム幅は半角文字数が単位のため、全角を2・半角を1として数えた
    文字数をカラム幅で割ることで折り返し後の行数を求める。
    セル内の左右パディング分として幅を1だけ差し引いて余裕を持たせる。
    """
    if not text:
        return 1
    capacity = max(1, col_width - 1)
    total_lines = 0
    for line in str(text).split("\n"):
        char_count = sum(2 if ord(c) > 127 else 1 for c in line)
        total_lines += max(1, -(-char_count // capacity))
    return max(1, total_lines)


def choose_font(category: str) -> Font:
    """カテゴリに応じたフォントを返す。"""
    if "【バグ修正確認】" in category:
        return STYLES["bug_fix"]["font"]
    if "【新規追加】" in category:
        return STYLES["new_feature"]["font"]
    return STYLES["data"]["font"]


def create_excel(data: dict, output_path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "テスト仕様書"

    num_cols = len(COLUMNS)

    # カラム幅
    for i, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # --- 1行目: タイトル ---
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=f"ロムスビ　結合テストチェックリスト　{data.get('title', '')}")
    cell.font = STYLES["title"]["font"]
    ws.row_dimensions[row].height = 28

    # --- 2行目: ブランチ ---
    row = 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=f"ブランチ: {data.get('branch', '')}")
    cell.font = STYLES["branch"]["font"]
    ws.row_dimensions[row].height = 22

    # --- 3行目: ヘッダー ---
    row = 3
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = STYLES["header"]["font"]
        cell.fill = STYLES["header"]["fill"]
        cell.alignment = STYLES["header"]["alignment"]
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 24

    # --- 4行目以降: セクション + テストケース ---
    row = 4
    data_row_idx = 0

    for section in data.get("sections", []):
        # セクション区切り行
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
        cell = ws.cell(row=row, column=1, value=f"■ {section.get('name', '')}")
        cell.font = STYLES["section"]["font"]
        cell.fill = STYLES["section"]["fill"]
        cell.alignment = Alignment(vertical="center")
        cell.border = THIN_BORDER
        for col_idx in range(2, num_cols + 1):
            c = ws.cell(row=row, column=col_idx)
            c.fill = STYLES["section"]["fill"]
            c.border = THIN_BORDER
        ws.row_dimensions[row].height = 26
        row += 1

        # テストケース
        for case in section.get("cases", []):
            data_row_idx += 1
            is_even = data_row_idx % 2 == 0
            category = str(case.get("category", ""))
            font = choose_font(category)

            values = [
                case.get("no", ""),
                category,
                case.get("test_item", ""),
                case.get("procedure", ""),
                case.get("expected", ""),
                "",  # 確認者１
                "",  # 確認者２
                case.get("notes", ""),
            ]

            max_lines = 1
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.font = font
                cell.alignment = STYLES["data"]["alignment"]
                cell.border = THIN_BORDER
                if is_even:
                    cell.fill = STYLES["data_even"]["fill"]
                # 全文が表示されるよう、最も行数が多いカラムに行高を合わせる
                max_lines = max(max_lines, count_wrapped_lines(str(value), COL_WIDTHS[col_idx - 1]))

            # No列は中央寄せ（折り返しは維持する）
            ws.cell(row=row, column=1).alignment = Alignment(
                horizontal="center", vertical="top", wrap_text=True
            )

            ws.row_dimensions[row].height = min(max_lines * LINE_HEIGHT, MAX_ROW_HEIGHT)
            row += 1

    # ウィンドウ枠固定（ヘッダー行まで）
    ws.freeze_panes = "A4"

    # 印刷設定
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1

    wb.save(output_path)
    return output_path


def main():
    if len(sys.argv) < 3:
        print(f"Usage: {sys.argv[0]} <input.json> <output.xlsx>", file=sys.stderr)
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2]

    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    result = create_excel(data, output_path)
    print(f"[OK] Excelファイルを生成しました: {result}")


if __name__ == "__main__":
    main()
